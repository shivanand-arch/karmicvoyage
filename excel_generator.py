"""
Excel report generation with conditional formatting, verdict colors, and summary sheet.
Matches the exact format used in Exotel screening sessions.
"""

import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def clean_str(s, max_len=300):
    if not isinstance(s, str):
        return s
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", s)[:max_len]


# ── Style constants ──────────────────────────

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

VERDICT_FILLS = {
    "Strong Yes": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "Yes": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
    "Maybe": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
    "No": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
}
VERDICT_FONTS = {
    "Strong Yes": Font(name="Arial", bold=True, color="FFFFFF", size=9),
    "Yes": Font(name="Arial", bold=True, color="FFFFFF", size=9),
    "Maybe": Font(name="Arial", bold=True, color="000000", size=9),
    "No": Font(name="Arial", bold=True, color="FFFFFF", size=9),
}
BONUS_HEADER_FILL = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")


def score_fill(val):
    if val >= 8:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val >= 6.5:
        return PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    elif val >= 5:
        return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    else:
        return PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def generate_excel(candidates: list, framework: dict, role_name: str) -> io.BytesIO:
    """
    Generate a formatted Excel workbook from candidate results.
    Returns a BytesIO buffer ready for download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rankings"

    dimensions = list(framework["dimensions"].keys())
    weights = framework["weights"]
    has_bonus = "bonus_dimensions" in framework
    bonus_dims = list(framework.get("bonus_dimensions", {}).keys())

    # ── Build headers ────────────────────────
    headers = ["Rank", "Name", "Current Role", "YOE"]
    for dim in dimensions:
        label = dim.replace("_", " ").title()
        weight_pct = f"{weights[dim]:.0%}"
        headers.append(f"{label}\n({weight_pct})")
    headers.extend(["Total\nScore", "Verdict"])

    # Bonus dimension headers
    for bd in bonus_dims:
        label = bd.replace("_", " ").title()
        headers.append(f"{label}*")
        headers.append(f"{label} Notes")

    headers.extend(["Key Strengths", "Key Concerns"])

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Mark bonus headers with purple
    base_cols = 4 + len(dimensions) + 2  # Rank,Name,Role,YOE + dims + Total,Verdict
    for i, bd in enumerate(bonus_dims):
        for offset in range(2):  # score and notes columns
            col_idx = base_cols + 1 + i * 2 + offset
            ws.cell(row=1, column=col_idx).fill = BONUS_HEADER_FILL

    # Column widths
    col_widths = [5, 24, 30, 5]
    col_widths += [11] * len(dimensions)
    col_widths += [9, 12]
    for bd in bonus_dims:
        col_widths += [9, 35]
    col_widths += [40, 40]

    for i, w in enumerate(col_widths, 1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 50
    ws.freeze_panes = "A2"

    data_font = Font(name="Arial", size=9)
    small_font = Font(name="Arial", size=8)

    # ── Write data rows ──────────────────────
    for idx, c in enumerate(candidates):
        row = idx + 2
        scores = c.get("scores", {})

        row_data = [
            idx + 1,
            clean_str(c.get("name", "Unknown")),
            clean_str(c.get("current_role", "Not specified")),
            c.get("yoe", 0),
        ]

        # Dimension scores
        for dim in dimensions:
            row_data.append(round(scores.get(dim, 0), 1))

        total = c.get("total_score", 0)
        verdict = c.get("verdict", "No")
        row_data.extend([total, verdict])

        # Bonus dimensions
        bonus_scores = c.get("bonus_scores", {})
        bonus_notes = c.get("bonus_notes", {})
        for bd in bonus_dims:
            row_data.append(bonus_scores.get(bd, 0))
            row_data.append(clean_str(bonus_notes.get(bd, "")))

        # Strengths and concerns
        strengths = c.get("key_strengths", [])
        concerns = c.get("key_concerns", [])
        row_data.append(clean_str(", ".join(strengths) if isinstance(strengths, list) else str(strengths)))
        row_data.append(clean_str(", ".join(concerns) if isinstance(concerns, list) else str(concerns)))

        # Write cells
        dim_start_col = 5
        dim_end_col = 4 + len(dimensions)
        total_col = dim_end_col + 1
        verdict_col = total_col + 1

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Score coloring for dimension columns
            if dim_start_col <= col <= dim_end_col and isinstance(val, (int, float)):
                cell.fill = score_fill(val)
                cell.number_format = "0.0"

            # Total score column
            if col == total_col:
                cell.fill = score_fill(val)
                cell.font = Font(name="Arial", bold=True, size=10)
                cell.number_format = "0.00"

            # Verdict column
            if col == verdict_col:
                cell.fill = VERDICT_FILLS.get(verdict, PatternFill())
                cell.font = VERDICT_FONTS.get(verdict, data_font)

            # Name and role — left align
            if col in (2, 3):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Bonus score columns
            bonus_start = verdict_col + 1
            for i, bd in enumerate(bonus_dims):
                score_col = bonus_start + i * 2
                note_col = score_col + 1
                if col == score_col and isinstance(val, (int, float)):
                    cell.fill = score_fill(val)
                    cell.font = Font(name="Arial", bold=True, size=10)
                    cell.number_format = "0.0"
                if col == note_col:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.font = small_font

            # Text columns at the end — left align, wrap, small font
            text_cols_start = bonus_start + len(bonus_dims) * 2
            if col >= text_cols_start:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = small_font

        ws.row_dimensions[row].height = 55

    # Auto-filter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(candidates) + 1}"

    # ── Summary sheet ────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    from collections import Counter
    vc = Counter(c["verdict"] for c in candidates)

    summary_rows = [
        ("Role", role_name),
        ("Company", "Exotel"),
        ("Total Resumes", len(candidates)),
        ("", ""),
        ("Verdict Distribution", "Count"),
        ("Strong Yes (≥8.0)", vc.get("Strong Yes", 0)),
        ("Yes (6.5-7.9)", vc.get("Yes", 0)),
        ("Maybe (5.0-6.4)", vc.get("Maybe", 0)),
        ("No (<5.0)", vc.get("No", 0)),
        ("", ""),
        ("Scoring Dimensions", "Weight"),
    ]
    for dim in dimensions:
        label = dim.replace("_", " ").title()
        summary_rows.append((label, f"{weights[dim]:.0%}"))

    if bonus_dims:
        summary_rows.append(("", ""))
        summary_rows.append(("* Bonus Dimensions", "(Separate)"))
        for bd in bonus_dims:
            label = bd.replace("_", " ").title()
            summary_rows.append((label, "Scored 1-10, not in total"))

    for r, (k, v) in enumerate(summary_rows, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(name="Arial", bold=bool(k), size=10)
        ws2.cell(row=r, column=2, value=v).font = Font(name="Arial", size=10)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
